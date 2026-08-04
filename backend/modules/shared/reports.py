"""
Smart Dairy ERP — Report Routes

GET  /api/reports?type=collection|payment|farmer|quality|rejection|branch|expense|pnl&from=&to=&branchId=
GET  /api/reports/export?type=...&format=csv — CSV export of any report
"""
from datetime import datetime, date, timedelta
import csv
import io
from flask import Blueprint, request, jsonify, Response
from backend.app import db
from backend.models import (
    Collection, Farmer, Payment, MilkRejection, QualityTest, Branch,
    Expense, VendorPayment, InventoryItem, PurchaseOrder, Vehicle, Employee,
)
from backend.auth import role_required, get_identity, get_branch_scope
from backend.audit import log_audit

report_bp = Blueprint('reports', __name__)

# Branch Managers may only request operational reports for their own branch.
# Company-wide / head-office reports (branch, pnl, procurement, inventory,
# vehicle, employee) stay restricted to global roles.
BM_ALLOWED_REPORT_TYPES = ('collection', 'payment', 'farmer', 'quality', 'rejection')


def _enforce_bm_report_access(report_type, branch_id):
    """
    Enforce report RBAC for Branch Managers.

    Returns (branch_id, error_or_None): branch_id is forced to the manager's
    own branch (any client-supplied branchId is ignored), and an error message
    is returned when the report type is not available to the role.
    """
    user = get_identity()
    if user.get('role') == 'BRANCH_MANAGER':
        if report_type not in BM_ALLOWED_REPORT_TYPES:
            return None, 'Access denied. This report is not available for your role.'
        # Fail-closed: unassigned branch managers are denied by get_branch_scope()
        return get_branch_scope(), None
    return branch_id, None


@report_bp.route('/api/reports', methods=['GET'])
@role_required('SUPER_ADMIN', 'HEAD_OFFICE', 'BRANCH_MANAGER')
def get_report():
    """Generate reports based on type and filters."""
    report_type = request.args.get('type', 'collection')
    from_date = request.args.get('from', '')
    to_date = request.args.get('to', '')
    branch_id = request.args.get('branchId', type=int)
    farmer_id = request.args.get('farmerId', type=int)

    # RBAC: Branch Managers are limited to operational reports and are always
    # scoped to their own branch (client-supplied branchId is ignored).
    branch_id, deny = _enforce_bm_report_access(report_type, branch_id)
    if deny:
        return jsonify({'error': deny}), 403

    # Parse dates
    start = None
    end = None
    try:
        if from_date:
            start = datetime.strptime(from_date, '%Y-%m-%d').date()
        if to_date:
            end = datetime.strptime(to_date, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        pass

    if not start:
        end = date.today()
        start = end - timedelta(days=30)

    if report_type == 'collection':
        return _collection_report(start, end, branch_id)
    elif report_type == 'payment':
        return _payment_report(start, end, branch_id)
    elif report_type == 'farmer':
        return _farmer_ledger(farmer_id, start, end)
    elif report_type == 'quality':
        return _quality_report(start, end, branch_id)
    elif report_type == 'rejection':
        return _rejection_report(start, end, branch_id)
    elif report_type == 'branch':
        return _branch_report(start, end)
    elif report_type == 'expense':
        return _expense_report(start, end, branch_id)
    elif report_type == 'pnl':
        return _pnl_report(start, end, branch_id)
    elif report_type == 'inventory':
        return _inventory_report(branch_id)
    elif report_type == 'procurement':
        return _procurement_report(start, end, branch_id)
    elif report_type == 'vehicle':
        return _vehicle_report()
    elif report_type == 'employee':
        return _employee_report(branch_id)
    else:
        return jsonify({'error': 'Invalid report type'}), 400


def _collection_report(start, end, branch_id):
    query = Collection.query.filter(Collection.date.between(start, end))
    if branch_id:
        query = query.filter_by(branch_id=branch_id)

    collections = query.all()

    total_qty = sum(c.quantity for c in collections)
    total_amount = sum(c.amount for c in collections)
    morning = sum(c.quantity for c in collections if c.shift == 'MORNING')
    evening = sum(c.quantity for c in collections if c.shift == 'EVENING')
    avg_fat = sum(c.fat for c in collections if c.fat) / max(len([c for c in collections if c.fat]), 1)
    avg_snf = sum(c.snf for c in collections if c.snf) / max(len([c for c in collections if c.snf]), 1)

    return jsonify({
        'type': 'collection',
        'period': {'from': start.isoformat(), 'to': end.isoformat()},
        'summary': {
            'totalQuantity': round(total_qty, 2),
            'totalAmount': round(total_amount, 2),
            'morningQuantity': round(morning, 2),
            'eveningQuantity': round(evening, 2),
            'collectionCount': len(collections),
            'avgFat': round(avg_fat, 2),
            'avgSnf': round(avg_snf, 2),
        },
        'collections': [c.to_dict() for c in collections[:100]],
    })


def _payment_report(start, end, branch_id):
    query = Payment.query.filter(Payment.created_at.between(
        datetime.combine(start, datetime.min.time()),
        datetime.combine(end, datetime.max.time())
    ))
    if branch_id:
        query = query.filter_by(branch_id=branch_id)

    payments = query.all()
    total_paid = sum(p.total_amount for p in payments if p.status == 'PAID')
    total_pending = sum(p.total_amount for p in payments if p.status in ('PENDING', 'APPROVED'))

    return jsonify({
        'type': 'payment',
        'period': {'from': start.isoformat(), 'to': end.isoformat()},
        'summary': {
            'totalPaid': round(total_paid, 2),
            'totalPending': round(total_pending, 2),
            'paymentCount': len(payments),
        },
        'payments': [p.to_dict() for p in payments[:100]],
    })


def _farmer_ledger(farmer_id, start, end):
    if not farmer_id:
        return jsonify({'error': 'Farmer ID is required for farmer ledger'}), 400

    farmer = Farmer.query.get(farmer_id)
    if not farmer:
        return jsonify({'error': 'Farmer not found'}), 404

    # Branch isolation: Branch Managers cannot view another branch's ledger
    forced = get_branch_scope()
    if forced and farmer.branch_id != forced:
        return jsonify({'error': 'Access denied. Farmer belongs to another branch.'}), 403

    collections = Collection.query.filter(
        Collection.farmer_id == farmer_id,
        Collection.date.between(start, end)
    ).order_by(Collection.date.desc()).all()

    total_qty = sum(c.quantity for c in collections)
    total_amount = sum(c.amount for c in collections)

    return jsonify({
        'type': 'farmer',
        'farmer': farmer.to_dict(),
        'period': {'from': start.isoformat(), 'to': end.isoformat()},
        'summary': {
            'totalQuantity': round(total_qty, 2),
            'totalAmount': round(total_amount, 2),
            'collectionCount': len(collections),
        },
        'collections': [c.to_dict() for c in collections],
    })


def _quality_report(start, end, branch_id):
    query = QualityTest.query.filter(QualityTest.date.between(start, end))
    if branch_id:
        query = query.filter_by(branch_id=branch_id)

    tests = query.all()
    passed = sum(1 for t in tests if t.overall_result == 'PASS')
    borderline = sum(1 for t in tests if t.overall_result == 'BORDERLINE')
    failed = sum(1 for t in tests if t.overall_result == 'FAIL')

    return jsonify({
        'type': 'quality',
        'period': {'from': start.isoformat(), 'to': end.isoformat()},
        'summary': {
            'totalTests': len(tests),
            'passed': passed,
            'borderline': borderline,
            'failed': failed,
            'passRate': round(passed / max(len(tests), 1) * 100, 1),
        },
    })


def _rejection_report(start, end, branch_id):
    query = MilkRejection.query.filter(MilkRejection.date.between(start, end))
    if branch_id:
        query = query.filter_by(branch_id=branch_id)

    rejections = query.all()
    total_qty = sum(r.quantity for r in rejections)

    # Group by reason
    reasons = {}
    for r in rejections:
        reasons[r.reason] = reasons.get(r.reason, 0) + r.quantity

    return jsonify({
        'type': 'rejection',
        'period': {'from': start.isoformat(), 'to': end.isoformat()},
        'summary': {
            'totalQuantity': round(total_qty, 2),
            'totalEvents': len(rejections),
            'byReason': reasons,
        },
    })


def _expense_report(start, end, branch_id):
    query = Expense.query.filter(Expense.expense_date.between(start, end))
    if branch_id:
        query = query.filter_by(branch_id=branch_id)

    expenses = query.all()
    total = sum(e.amount or 0 for e in expenses)
    by_category = {}
    for e in expenses:
        by_category[e.category] = round(by_category.get(e.category, 0) + (e.amount or 0), 2)

    return jsonify({
        'type': 'expense',
        'period': {'from': start.isoformat(), 'to': end.isoformat()},
        'summary': {
            'totalAmount': round(total, 2),
            'expenseCount': len(expenses),
            'byCategory': by_category,
        },
        'expenses': [e.to_dict() for e in expenses[:200]],
    })


def _pnl_report(start, end, branch_id):
    """Profit & Loss: revenue (milk) − expenses − procurement spend."""
    coll_query = Collection.query.filter(Collection.date.between(start, end))
    if branch_id:
        coll_query = coll_query.filter_by(branch_id=branch_id)
    collections = coll_query.all()
    revenue = round(sum(c.amount or 0 for c in collections), 2)
    total_qty = round(sum(c.quantity or 0 for c in collections), 2)

    exp_query = Expense.query.filter(Expense.expense_date.between(start, end))
    if branch_id:
        exp_query = exp_query.filter_by(branch_id=branch_id)
    expenses = exp_query.all()
    total_expenses = round(sum(e.amount or 0 for e in expenses), 2)

    vp_query = VendorPayment.query.filter(VendorPayment.payment_date.between(start, end))
    procurement_spend = round(sum(p.amount or 0 for p in vp_query.all()), 2)

    # Farmer payments (paid) are the cost of milk to the dairy
    paid_query = Payment.query.filter(
        Payment.status == 'PAID',
        Payment.paid_at.isnot(None),
    )
    if branch_id:
        paid_query = paid_query.filter_by(branch_id=branch_id)
    farmer_payments = round(sum(p.total_amount or 0 for p in paid_query.all()), 2)

    total_costs = round(total_expenses + procurement_spend + farmer_payments, 2)
    net = round(revenue - total_costs, 2)

    return jsonify({
        'type': 'pnl',
        'period': {'from': start.isoformat(), 'to': end.isoformat()},
        'summary': {
            'revenue': revenue,
            'totalQuantity': total_qty,
            'expenses': total_expenses,
            'procurementSpend': procurement_spend,
            'farmerPayments': farmer_payments,
            'totalCosts': total_costs,
            'profit': net if net >= 0 else 0,
            'loss': abs(net) if net < 0 else 0,
            'net': net,
        },
    })


def _inventory_report(branch_id=None):
    items = InventoryItem.query.all()
    if branch_id:
        items = [i for i in items if i.branch_id == branch_id]

    low = [i for i in items if i.stock <= (i.min_stock or 0)]
    return jsonify({
        'type': 'inventory',
        'summary': {
            'totalItems': len(items),
            'totalStock': round(sum(i.stock or 0 for i in items), 2),
            'lowStockCount': len(low),
        },
        'items': [i.to_dict() for i in items],
    })


def _procurement_report(start, end, branch_id):
    query = PurchaseOrder.query.filter(PurchaseOrder.order_date.between(start, end))
    if branch_id:
        query = query.filter_by(branch_id=branch_id)
    pos = query.all()

    total_value = sum(p.total_amount or 0 for p in pos)
    total_paid = sum(p.paid_amount or 0 for p in pos)
    status_counts = {}
    for p in pos:
        status_counts[p.status] = status_counts.get(p.status, 0) + 1

    return jsonify({
        'type': 'procurement',
        'period': {'from': start.isoformat(), 'to': end.isoformat()},
        'summary': {
            'poCount': len(pos),
            'totalValue': round(total_value, 2),
            'totalPaid': round(total_paid, 2),
            'byStatus': status_counts,
        },
        'purchase_orders': [p.to_dict() for p in pos[:100]],
    })


def _vehicle_report():
    vehicles = Vehicle.query.all()
    from datetime import date, timedelta
    today = date.today()
    due_soon = today + timedelta(days=14)
    service_due = [v for v in vehicles if v.next_service_date and today <= v.next_service_date <= due_soon]
    insurance_due = [v for v in vehicles if v.insurance_expiry and today <= v.insurance_expiry <= due_soon]
    return jsonify({
        'type': 'vehicle',
        'summary': {
            'totalVehicles': len(vehicles),
            'active': sum(1 for v in vehicles if v.status == 'ACTIVE'),
            'maintenance': sum(1 for v in vehicles if v.status == 'MAINTENANCE'),
            'serviceDue': len(service_due),
            'insuranceExpiring': len(insurance_due),
        },
        'vehicles': [v.to_dict() for v in vehicles],
    })


def _employee_report(branch_id=None):
    query = Employee.query
    if branch_id:
        query = query.filter_by(branch_id=branch_id)
    employees = query.all()
    total_salary = sum(e.salary or 0 for e in employees)
    roles = {}
    for e in employees:
        roles[e.role] = roles.get(e.role, 0) + 1
    return jsonify({
        'type': 'employee',
        'summary': {
            'totalEmployees': len(employees),
            'active': sum(1 for e in employees if e.status == 'ACTIVE'),
            'totalSalary': round(total_salary, 2),
            'byRole': roles,
        },
        'employees': [e.to_dict() for e in employees],
    })


def _flatten_report(data, report_type):
    """Convert a report payload into CSV rows (headers + rows)."""
    summary = data.get('summary', {})
    rows = []

    def add(header, value):
        rows.append((header, value))

    add('Report Type', report_type)
    period = data.get('period', {})
    add('From', period.get('from', ''))
    add('To', period.get('to', ''))
    for key, value in summary.items():
        if isinstance(value, dict):
            for k, v in value.items():
                add(f'{key}.{k}', v)
        else:
            add(key.replace('_', ' ').title(), value)

    return rows


def _build_csv(header_pairs, rows, filename):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(header_pairs)
    writer.writerow([])
    for row in rows:
        writer.writerow(row)
    return Response(
        '\ufeff' + output.getvalue(),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


@report_bp.route('/api/reports/export', methods=['GET'])
@role_required('SUPER_ADMIN', 'HEAD_OFFICE', 'BRANCH_MANAGER')
def export_report():
    """Export any report as CSV."""
    report_type = request.args.get('type', 'collection')
    from_date = request.args.get('from', '')
    to_date = request.args.get('to', '')
    branch_id = request.args.get('branchId', type=int)
    farmer_id = request.args.get('farmerId', type=int)
    fmt = request.args.get('format', 'csv').lower()
    if fmt not in ('csv', 'xlsx', 'pdf'):
        return jsonify({'error': "Format must be csv, xlsx or pdf"}), 400

    # RBAC: same restrictions as the report endpoint
    branch_id, deny = _enforce_bm_report_access(report_type, branch_id)
    if deny:
        return jsonify({'error': deny}), 403

    # Reuse the same report generation logic
    start, end = None, None
    try:
        if from_date:
            start = datetime.strptime(from_date, '%Y-%m-%d').date()
        if to_date:
            end = datetime.strptime(to_date, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        pass
    if not start:
        end = date.today()
        start = end - timedelta(days=30)

    handlers = {
        'collection': _collection_report,
        'payment': _payment_report,
        'quality': _quality_report,
        'rejection': _rejection_report,
        'branch': _branch_report,
        'expense': _expense_report,
        'pnl': _pnl_report,
        'inventory': lambda s, e, b: _inventory_report(b),
        'procurement': _procurement_report,
        'vehicle': lambda s, e, b: _vehicle_report(),
        'employee': lambda s, e, b: _employee_report(b),
    }
    if report_type == 'farmer':
        return jsonify({'error': 'Use the farmer export endpoint for farmer statements'}), 400
    if report_type not in handlers:
        return jsonify({'error': 'Invalid report type'}), 400

    result = handlers[report_type](start, end, branch_id)
    data = result.get_json()

    log_audit('EXPORT', f'{report_type.capitalize()}Report', None,
              detail=f'Exported {report_type} report as {fmt.upper()}')
    db.session.commit()

    filename = f'{report_type}_report_{date.today().isoformat()}'
    if fmt == 'csv':
        rows = _flatten_report(data, report_type)
        if report_type == 'branch':
            branch_rows = []
            for b in data.get('branches', []):
                branch_rows.append([
                    b.get('branchCode', ''), b.get('branchName', ''),
                    b.get('farmerCount', 0), b.get('totalQuantity', 0),
                    b.get('totalAmount', 0), b.get('collectionCount', 0),
                ])
            rows.extend([['Branch', 'Name', 'Farmers', 'Quantity', 'Amount', 'Collections']])
            rows.extend(branch_rows)
        return _build_csv(rows, [], filename + '.csv')
    elif fmt == 'xlsx':
        return _build_xlsx(data, report_type, filename + '.xlsx')
    elif fmt == 'pdf':
        return _build_pdf(data, report_type, filename + '.pdf')
    else:
        return jsonify({'error': "Format must be csv, xlsx or pdf"}), 400


def _build_xlsx(data, report_type, filename):
    """Export report summary as an Excel workbook (openpyxl)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = f'{report_type.title()} Report'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')

    # Header row: report title + period
    ws.append([f'Smart Dairy ERP — {report_type.title()} Report'])
    ws['A1'].font = Font(bold=True, size=14)
    period = data.get('period', {})
    ws.append([f'Period: {period.get("from", "")} to {period.get("to", "")}'])
    ws.append([])

    # Summary section
    summary = data.get('summary', {})
    ws.append(['SUMMARY'])
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill
    for key, value in summary.items():
        if isinstance(value, dict):
            for k, v in value.items():
                ws.append([f'{key}.{k}', v])
        else:
            ws.append([key.replace('_', ' ').title(), value])
    ws.append([])

    # Detail section (list of records if present)
    for detail_key in ('items', 'vehicles', 'employees', 'collections', 'payments',
                       'expenses', 'purchase_orders', 'branches', 'rejections', 'tests'):
        records = data.get(detail_key)
        if not records:
            continue
        ws.append([detail_key.upper()])
        header_row = list(records[0].keys())
        ws.append(header_row)
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
        for rec in records:
            ws.append([rec.get(k, '') for k in header_row])
        break

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


def _build_pdf(data, report_type, filename):
    """Export report summary as a PDF (reportlab)."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, title=f'Smart Dairy ERP - {report_type} Report')
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('TitleX', parent=styles['Title'], fontSize=16, textColor=colors.HexColor('#1B4332'))
    elements = []
    elements.append(Paragraph(f'Smart Dairy ERP — {report_type.title()} Report', title_style))
    period = data.get('period', {})
    elements.append(Paragraph(f'Period: {period.get("from", "")} to {period.get("to", "")}', styles['Normal']))
    elements.append(Spacer(1, 12))

    rows = _flatten_report(data, report_type)
    table = Table([[r[0], r[1]] for r in rows])
    table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F0E8')),
        ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(f'Generated: {datetime.utcnow().strftime("%d %b %Y %H:%M")} UTC', styles['Normal']))
    doc.build(elements)
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment; filename={filename}'},
    )


def _branch_report(start, end):
    branches = Branch.query.filter_by(status='ACTIVE').all()
    branch_data = []

    for branch in branches:
        collections = Collection.query.filter(
            Collection.branch_id == branch.id,
            Collection.date.between(start, end),
        ).all()

        total_qty = sum(c.quantity for c in collections)
        total_amt = sum(c.amount for c in collections)
        farmer_count = Farmer.query.filter_by(branch_id=branch.id, status='ACTIVE').count()

        branch_data.append({
            'branchId': branch.id,
            'branchName': branch.name,
            'branchCode': branch.code,
            'farmerCount': farmer_count,
            'totalQuantity': round(total_qty, 2),
            'totalAmount': round(total_amt, 2),
            'collectionCount': len(collections),
        })

    return jsonify({
        'type': 'branch',
        'period': {'from': start.isoformat(), 'to': end.isoformat()},
        'branches': branch_data,
    })
